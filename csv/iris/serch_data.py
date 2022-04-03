import openpyxl
import glob
import re

def conve_cell(rowcol):
    pattern1='.*?(\d+).*'
    pattern2="[A-Z]+"
    
    rnum=re.findall(pattern1,rowcol)
    alpha=re.findall(pattern2,rowcol,flags=re.IGNORECASE)
    
    alpha=alpha[0].upper()
    def alpha2num(alpha):
        cnum=0
        for index, item in enumerate(list(alpha)):
            cnum += pow(26,len(alpha)-index-1)*(ord(item)-ord('A')+1)
        return cnum

    return int(rnum[0]),alpha2num(alpha)

celllist=[]   
wbori=r"C:\Users\qzkl2\Documents\データ取得2.xlsx"
sheetori="データ貼り付け"
wb=openpyxl.load_workbook(wbori)
sheet=wb[sheetori]    

oricell=[1,4]

i=0
while sheet.cell(row=oricell[0],column=oricell[1]+i).value is not None:
    concell=sheet.cell(row=oricell[0],column=oricell[1]+i).value
    celllist.append(conve_cell(concell))
    i+=1

  
    
