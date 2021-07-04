import openpyxl as px

wb='weather.xlsx'
wb2='weather2.xlsx'

wbw=px.load_workbook(wb, data_only=True)
print(wbw.sheetnames)

wsw=wbw.worksheets[0]
wsw.title="SheetOne"
print(wsw.title)

# for row in wsw.values:
#     for value in row:
#         print(value)

for row in wsw.iter_rows(min_row=1):
    
    for col in row:
        print(col.value)
    

wbw.save(wb2)
