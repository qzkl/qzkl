import time
import pathlib
import openpyxl as px
import csv
import prime_num

wb = px.Workbook()
ws = wb.active


def fib(a):
    if a==1:
        return 1
    elif a==2:
        return 1
    else:
        return fib(a-1)+fib(a-2)

def fib2(n,a=1,b=0):
    
    
    if n==1:
        return a
    
    else:
        return fib2(n-1,a+b,a)

n=1
for i in range(1,100):

    start = time.time()
    print(i)
    x=fib2(i)

    end = time.time()
    print(f'処理にかかった平均時間は{(end-start)}秒です。')
    ws.cell(row = n, column = 1, value = i)
    ws.cell(row = n, column = 2, value = x)
    
    ws.cell(row = n, column = 3, value = (end-start))
    n+=1



# for i in range(0,10):
#     ws.cell(row = 1, column = i+1, value = p_num[i])


ps="C:/bbpy/vs/"
fname='fib2.xlsx'
sfile=ps+fname
wb.save(sfile)
