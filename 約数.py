import prime_num
import pathlib
import openpyxl as px
import csv


p_num=prime_num.prime_n(2000)

wb = px.Workbook()
ws = wb.active


for i in range(0,len(p_num)):
    ws.cell(row = 1, column = i+2, value = p_num[i])

def mod_e(id,b,by,i,num_e=0):
    
    mod=id%by
    e=id//by
    if mod==0 :
        num_e+=1#約数の数
        
        return mod_e(e,b,by,i,num_e)
    else:
        # print("{}:{}".format(by,num_e))
        
        if num_e != 0:
            s=(by**(num_e+1)-1)/(by-1)
            ws.cell(row = b, column = i+2, value = s)
        
        return id

def div(m,b,n=0,i=0):
    # print("数：{}".format(m))
    if m==1:
        pass
    else:
        e=mod_e(m,b,p_num[n],n)
        n+=1
        return div(e,b,n)

cnt=2
# f=div(128,cnt)
for f_num in range(1,1001):
    
    ws.cell(row = cnt, column = 1, value = f_num)
    f=div(f_num,cnt)
    cnt+=1



ps="C:/bbpy/vs/エクセル/"
fname='sample.xlsx'
sfile=ps+fname
wb.save(sfile)
